import openpyxl
import sys

def get_employers(name1, tab_keyword, fio_keyword, role1_keywords, role2_keywords, kpi_keywords):

    # Open first file
    wb = openpyxl.load_workbook(name1)
    sheetnames = list(map(lambda x: x.title, wb.worksheets))
    sheetindexes = list(map(lambda x: str(sheetnames.index(x) + 1), sheetnames))
    sheetnumber = 0

    # If more than one sheet in first file
    if len(wb.worksheets) > 1:
        print(f'В документе {name1} несколько листов:')
        for sheetname in sheetnames:
            print(f'{sheetnames.index(sheetname) + 1}: {sheetname}')
        
        stay_in_loop = True
        quitchoice = ['quit']
        loopindex = 0
        while stay_in_loop:
            print(f'Введите номер нужного листа и нажмите enter:')
            if loopindex != 0:
                print('(чтобы отменить выполнение, введите quit)')
            choice = input()
            if choice in quitchoice:
                sys.exit('\nРабота скрипта завершена.')
            if choice not in sheetindexes:
                print('\nТакого номера нет!')
            else:
                sheetnumber = int(choice) - 1
                stay_in_loop = False
            loopindex += 1

    sheet = wb.worksheets[sheetnumber]

    # 
    tab_row = 0

    tab_column = 0
    fio_column = 0
    role1_column = 0
    role2_column = 0
    kpi_column = 0

    kpi_columns = []
    kpi_titles = []

    # Search for title row
    for row in sheet.rows:
        for cell in row:
            if tab_keyword in str(cell.value).lower():
                tab_column = cell.column
                tab_row = cell.row
                break
        else:
            continue
        break
        
    # Search for columns, that are needed
    for cell in list(sheet.rows)[tab_row - 1]:
        if fio_keyword.lower() in str(cell.value).lower() and fio_column == 0:
            fio_column = cell.column
        
        role1_flag = True
        for role1_keyword in role1_keywords:
            if role1_keyword.lower() not in str(cell.value).lower():
                role1_flag = False
        if role1_flag:
            role1_column = cell.column

        role2_flag = True
        for role2_keyword in role2_keywords:
            if role2_keyword.lower() not in str(cell.value).lower():
                role2_flag = False
        if role2_flag:
            role2_column = cell.column

        for kpi_keyword in kpi_keywords:
            if kpi_keyword.lower() in str(cell.value).lower():
                kpi_columns.append(cell.column)
                kpi_titles.append(cell.value)
                break

    # If data not found
    if tab_column == 0:
        sys.exit(f'\nCкрипт не нашел колонку с ключевым словом «{tab_keyword}» и завершил работу.')
    if fio_column == 0:
        sys.exit(f'\nCкрипт не нашел колонку с ключевым словом «{fio_keyword}» и завершил работу.')
    if role1_column == 0:
        sys.exit(f'\nCкрипт не нашел колонку с ключевыми словами {role1_keywords} и завершил работу.')
    if role2_column == 0:
        sys.exit(f'\nCкрипт не нашел колонку с ключевыми словами {role2_keywords} и завершил работу.')
    if not kpi_columns:
        sys.exit(f'\nCкрипт не нашел колонок с хотя бы одним из ключевых слов {kpi_keywords} и завершил работу.')

    # If there are many output (kpi) columns, user selects one
    if len(kpi_columns) > 1:

        kpi_indexes = list(map(lambda x: str(kpi_titles.index(x) + 1), kpi_titles))
        stay_in_loop = True
        quitchoice = ['quit']
        loopindex = 0

        print(f'\nВ документе {name1} несколько колонок с ключевыми словами {kpi_keywords}:')
        for kpi_title in kpi_titles:
            print(f'{kpi_titles.index(kpi_title) + 1}: {kpi_title}')

        while stay_in_loop:
            print('Введите номер нужной колонки и нажмите enter:')
            if loopindex != 0:
                print('(чтобы отменить выполнение, введите quit)')
            choice = input()
            if choice in quitchoice:
                sys.exit('\nРабота скрипта завершена.')
            if choice not in kpi_indexes:
                print('\nТакого номера нет!')
            else:
                kpi_column = kpi_columns[int(choice) - 1]
                stay_in_loop = False
            loopindex += 1
        
        # print(kpi_column)

    else:
        kpi_column = kpi_columns[0]

    #
    table = []

    for row in sheet.rows:
        index = list(sheet.rows).index(row) + 1
        table_row = []
        if list(row)[tab_column-1].value != None and index > tab_row:
            id = list(row)[tab_column-1].value
            fio = list(row)[fio_column-1].value
            role1 = list(row)[role1_column-1].value
            role2 = list(row)[role2_column-1].value
            table_row.append(id)
            table_row.append(fio)
            table_row.append(role1)
            table_row.append(role2)
            table_row.append('')
            table.append(table_row)
            # print(list(row)[tab_column-1].value)
    
    return table, sheetnumber, kpi_column, tab_row, tab_column

def find_kpis(name2, table, tab_keywords, intp_keywords):
    # Open second file
    wb = openpyxl.load_workbook(name2)
    # sheetnames = list(map(lambda x: x.title, wb.worksheets))
    # sheetindexes = list(map(lambda x: str(sheetnames.index(x) + 1), sheetnames))
    ids_table = [x[0] for x in table]
    
    tab_column = 0
    intp_column = 0

    for sheet in wb.worksheets:
        print(f'\n{sheet.title}')

        for row in sheet.rows:
            for cell in row:
                if cell.value != None:
                    for tab_keyword in tab_keywords:
                        if tab_keyword.lower() in str(cell.value).lower():
                            tab_column = cell.column
                            break
                    else:
                        continue
                    break
            else:
                continue
            break

        for row in sheet.rows:
            for cell in row:
                if cell.value != None:
                    for intp_keyword in intp_keywords:
                        if intp_keyword.lower() in str(cell.value).lower():
                            intp_column = cell.column
                            break
                    else:
                        continue
                    break
            else:
                continue
            break

        # If data not found
        if tab_column == 0:
            print(f'{sheet.title}: не найдено ключевых слов {tab_keywords}')
            continue
        if intp_column == 0:
            print(f'{sheet.title}: не найдено ключевых слов {intp_keywords}')
            continue

        # print(tab_column, intp_column)

        for row in sheet.rows:
            possible_id = list(row)[tab_column-1].value
            if possible_id in ids_table:
                index = ids_table.index(possible_id)
                possible_intp = list(row)[intp_column-1].value
                table[index][-1] = possible_intp
    
    return table
            
def write_back(name1, table, sheetnumber, kpi_column, tab_row, tab_column):
    # print(table)
    # Open first file
    wb = openpyxl.load_workbook(name1)

    sheet = wb.worksheets[sheetnumber]
    k = 0
    for row in sheet.rows:
        index = list(sheet.rows).index(row) + 1
        if list(row)[tab_column-1].value != None and index > tab_row:
            
            sheet.cell(row=index, column=kpi_column).value = table[k][-1]
            k += 1
    wb.save(name1)

def main():
    name1 = 'file1.xlsx'
    name2 = 'file2.xlsx'

    # keywords for file 1
    tab_keyword = 'таб'
    fio_keyword = 'фио'
    role1_keywords = ['роль', 'мб']
    role2_keywords = ['роль', 'тд']
    kpi_keywords = ['kpi', 'кпи']

    # keywords for file 2
    tab_keywords_f2 = ['таб', 'т.н.']
    intp_keywords_f2 = ['интегр']

    # Start instructions
    print(f'\nПривет!\nУбедитесь, что исходные файлы {name1} и {name2} находятся в одной папке с исполняемым скриптом.\nНажмите enter, чтобы продолжить.')
    x = input()
    if x:
        print('')

    # Get info from employers list
    table, sheetnumber, kpi_column, tab_row, tab_column = get_employers(name1, tab_keyword, fio_keyword, role1_keywords, role2_keywords, kpi_keywords)
    # print(table)
    # print(kpi_column)

    # Get integral coefficients
    table = find_kpis(name2, table, tab_keywords_f2, intp_keywords_f2)

    # Write kpi back to file 1
    write_back(name1, table, sheetnumber, kpi_column, tab_row, tab_column)

    # Finish
    print('\nСкрипт закончил работу.')

if __name__ == '__main__':
    main()